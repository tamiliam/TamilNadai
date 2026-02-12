"""Generate Excel review file from Tier 2 and Tier 3 JSONL data."""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

sys.stdout.reconfigure(encoding="utf-8")


def style_header(ws, row, num_cols):
    """Apply header styling."""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_cell(cell, wrap=True):
    """Basic cell styling."""
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def add_tier2_sheet(wb, records):
    """Add Tier 2 review sheet."""
    ws = wb.active
    ws.title = "Tier 2 — Error Pairs"

    # Headers
    headers = ["#", "Rule ID", "Category", "Rule Name", "பிழை (Error)", "திருத்தம் (Correct)", "OK?", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    # Data rows
    for i, r in enumerate(records):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=r.get("rule_id", ""))
        ws.cell(row=row, column=3, value=r.get("category", ""))
        ws.cell(row=row, column=4, value=r.get("rule_name", ""))
        ws.cell(row=row, column=5, value=r.get("error_sentence", ""))
        ws.cell(row=row, column=6, value=r.get("correct_sentence", ""))
        ws.cell(row=row, column=7, value="")  # OK? column for reviewer
        ws.cell(row=row, column=8, value="")  # Notes column for reviewer

        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 30

    # Freeze header row
    ws.freeze_panes = "A2"

    return ws


def add_tier3_sheet(wb, records):
    """Add Tier 3 review sheet."""
    ws = wb.create_sheet("Tier 3 — Correct Sentences")

    headers = ["#", "Category", "Correct Sentence", "Note", "OK?", "Correction (if wrong)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    for i, r in enumerate(records):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=r.get("subcategory", r.get("category_tested", "")))
        ws.cell(row=row, column=3, value=r.get("correct_sentence", ""))
        ws.cell(row=row, column=4, value=r.get("short_description", r.get("note", "")))
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value="")

        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row, column=col))

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 45

    ws.freeze_panes = "A2"

    return ws


def main():
    project_dir = Path(__file__).resolve().parent.parent
    dataset_dir = project_dir / "dataset"
    review_dir = project_dir / "review"
    review_dir.mkdir(exist_ok=True)

    # Load data
    tier2_path = dataset_dir / "tier2_book_rules.jsonl"
    tier3_path = dataset_dir / "tier3_correct.jsonl"

    with open(tier2_path, "r", encoding="utf-8") as f:
        tier2 = [json.loads(line) for line in f if line.strip()]

    with open(tier3_path, "r", encoding="utf-8") as f:
        tier3 = [json.loads(line) for line in f if line.strip()]

    # Build Excel
    wb = Workbook()
    add_tier2_sheet(wb, tier2)
    add_tier3_sheet(wb, tier3)

    # Add instructions sheet
    ws_info = wb.create_sheet("Instructions", 0)
    instructions = [
        ["TamilNadai Dataset — Expert Review"],
        [""],
        ["Sheet 1: Tier 2 — Error Pairs (106 examples)"],
        ["  - Each row has an error sentence and its correction"],
        ["  - Mark 'OK' column with \u2713 if the pair is correct"],
        ["  - If wrong, write the correct form in Notes"],
        [""],
        ["Sheet 2: Tier 3 — Correct Sentences (100 examples)"],
        ["  - Each sentence should have NO grammar errors"],
        ["  - Mark 'OK' column with \u2713 if the sentence is correct"],
        ["  - If it has an error, write the correction in the last column"],
        [""],
        ["Source: \u0ba4\u0bae\u0bbf\u0bb4\u0bcd \u0ba8\u0b9f\u0bc8\u0b95\u0bcd \u0b95\u0bc8\u0baf\u0bc7\u0b9f\u0bc1 (Tamil Writing Style Guide)"],
        ["Publisher: Tamil Virtual Academy (tamilvu.org)"],
    ]
    for i, row_data in enumerate(instructions):
        cell = ws_info.cell(row=i + 1, column=1, value=row_data[0] if row_data else "")
        if i == 0:
            cell.font = Font(bold=True, size=14)
        elif row_data and row_data[0].startswith("Sheet"):
            cell.font = Font(bold=True, size=11)
    ws_info.column_dimensions["A"].width = 70

    out_path = review_dir / "TamilNadai_Review.xlsx"
    wb.save(out_path)
    print(f"Review file: {out_path}")
    print(f"  Tier 2: {len(tier2)} error pairs")
    print(f"  Tier 3: {len(tier3)} correct sentences")
    print(f"  Total: {len(tier2) + len(tier3)} items to review")


if __name__ == "__main__":
    main()
