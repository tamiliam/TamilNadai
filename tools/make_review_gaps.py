"""Generate Excel review file for the 41 gap-filling examples."""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

sys.stdout.reconfigure(encoding="utf-8")


def main():
    project_dir = Path(__file__).resolve().parent.parent
    dataset_dir = project_dir / "dataset"
    review_dir = project_dir / "review"

    jsonl_path = dataset_dir / "tier2_primary_gaps.jsonl"
    with open(jsonl_path, "r", encoding="utf-8") as f:
        records = [json.loads(line) for line in f if line.strip()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Gap Examples — Review"

    headers = ["#", "Rule ID", "Page", "Category", "Rule Name",
               "பிழை (Error)", "திருத்தம் (Correct)", "OK?", "Notes"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, r in enumerate(records):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=r.get("rule_id", ""))
        ws.cell(row=row, column=3, value=r.get("book_page", ""))
        ws.cell(row=row, column=4, value=r.get("category", ""))
        ws.cell(row=row, column=5, value=r.get("rule_name", ""))
        if r.get("is_error_example", True):
            ws.cell(row=row, column=6, value=r.get("error_sentence", ""))
        else:
            ws.cell(row=row, column=6, value="(correct — no error)")
        ws.cell(row=row, column=7, value=r.get("correct_sentence", ""))
        ws.cell(row=row, column=8, value="")
        ws.cell(row=row, column=9, value=r.get("short_description", ""))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 45
    ws.column_dimensions["H"].width = 8
    ws.column_dimensions["I"].width = 40
    ws.freeze_panes = "A2"

    out_path = review_dir / "TamilNadai_Gaps_Review.xlsx"
    wb.save(out_path)
    print(f"Review file: {out_path}")
    print(f"  {len(records)} examples to review")


if __name__ == "__main__":
    main()
